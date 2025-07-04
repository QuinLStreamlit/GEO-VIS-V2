import itertools
from matplotlib.ticker import FuncFormatter
from scipy.interpolate import make_interp_spline # Keep import for optional use
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns # For KDE plots (optional)
import os
import warnings
from pandas.api.types import is_numeric_dtype, is_object_dtype
from matplotlib.container import BarContainer # NEW IMPORT
from typing import List, Optional, Union, Sequence, Dict, Tuple, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re


def Excel_writer(df, filename, no_of_id_column=7):
    """
    Creates formatted Excel file with color coding and proper alignment.

    Features:
    - Header: Bold, size 11, blue background, dynamic height with text wrapping
    - First 3 columns: Left-aligned, light blue background (identifiers)
    - Question columns (contain '?'): Bold, yellow background, center-aligned
    - Regular columns: Center-aligned
    - Professional borders and auto-sizing
    """

    # Create workbook and add data
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define borders
    medium_border = Border(
        left=Side(style='medium', color='4472C4'),
        right=Side(style='medium', color='4472C4'),
        top=Side(style='medium', color='4472C4'),
        bottom=Side(style='medium', color='4472C4')
    )
    thin_border = Border(
        left=Side(style='thin', color='B8CCE4'),
        right=Side(style='thin', color='B8CCE4'),
        top=Side(style='thin', color='B8CCE4'),
        bottom=Side(style='thin', color='B8CCE4')
    )

    # Define colors
    colors = {
        'header_bg': '4472C4',  # Professional blue
        'header_text': 'FFFFFF',  # White
        'identifier_bg': 'E8F1FF',  # Light blue
        'identifier_text': '1F4788',  # Dark blue
        'question_bg': 'FFF2CC',  # Light yellow
        'question_text': '7F6000',  # Dark brown
        'alternate_row': 'F8F9FA'  # Light gray
    }

    # Define styles
    header_font = Font(bold=True, size=11, color=colors['header_text'])
    header_fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")

    identifier_font = Font(color=colors['identifier_text'])
    identifier_fill = PatternFill(start_color=colors['identifier_bg'], end_color=colors['identifier_bg'],
                                  fill_type="solid")

    question_font = Font(bold=True, color=colors['question_text'])
    question_fill = PatternFill(start_color=colors['question_bg'], end_color=colors['question_bg'], fill_type="solid")

    alternate_fill = PatternFill(start_color=colors['alternate_row'], end_color=colors['alternate_row'],
                                 fill_type="solid")

    # Define alignments with text wrapping for headers
    header_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_left = Alignment(horizontal="left", vertical="center")
    data_center = Alignment(horizontal="center", vertical="center")

    # Calculate dynamic header height based on longest column name
    max_header_length = max(len(str(col)) for col in df.columns)
    # Base height of 25px + additional space for every 20 characters
    header_height = max(30, 25 + (max_header_length // 20) * 15)
    ws.row_dimensions[1].height = header_height

    # Format each column
    for col in range(1, len(df.columns) + 1):
        column_header = df.columns[col - 1]
        column_letter = ws.cell(row=1, column=col).column_letter

        # Format header cell
        header_cell = ws.cell(row=1, column=col)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.border = medium_border

        # Determine column type
        is_identifier = col <= no_of_id_column
        is_question = re.search(r'\?', column_header)

        # Set header alignment
        header_cell.alignment = header_left if is_identifier else header_center

        # Calculate max length for auto-fitting column width
        max_length = 0
        for row in range(1, len(df) + 2):  # Iterate through all rows including header
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length

        # Auto-fit column width by setting it to the max length of its content plus a buffer.
        ws.column_dimensions[column_letter].width = max_length + 2

        # Format data cells
        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            if is_identifier:
                cell.fill = identifier_fill
                cell.font = identifier_font
                cell.alignment = data_left
            elif is_question:
                cell.fill = question_fill
                cell.font = question_font
                cell.alignment = data_center
            else:
                if row % 2 == 0:  # Alternating rows
                    cell.fill = alternate_fill
                cell.alignment = data_center

    wb.save(filename)

